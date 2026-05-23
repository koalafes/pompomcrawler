from __future__ import annotations

import csv
from pathlib import Path

from .extract import merge_duplicates
from .html_calendar import export_calendar_html
from .models import SCHEDULE_COLUMNS, ScheduleItem
from .schedule_window import DEFAULT_PAST_DAYS, filter_schedule_window


def export_schedule(
    items: list[ScheduleItem],
    output_dir: Path = Path("outputs"),
    *,
    filter_window: bool = True,
    past_days: int = DEFAULT_PAST_DAYS,
) -> tuple[Path, Path | None, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    merged = merge_duplicates(items)
    if filter_window:
        merged = filter_schedule_window(merged, past_days=past_days)
    csv_path = output_dir / "pompompurin_schedule.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=SCHEDULE_COLUMNS)
        writer.writeheader()
        for item in merged:
            writer.writerow(item.to_dict())

    xlsx_path = output_dir / "pompompurin_schedule.xlsx"
    html_path = export_calendar_html(merged, output_dir, filter_window=filter_window, past_days=past_days)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return csv_path, None, html_path

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "schedule"
    sheet.append(SCHEDULE_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EAD7A4")
    for item in merged:
        sheet.append([item.to_dict().get(column, "") for column in SCHEDULE_COLUMNS])
    for index, column in enumerate(SCHEDULE_COLUMNS, start=1):
        width = max(12, min(60, len(column) + 4))
        if column in {"title", "source_url", "image_url", "review_reason", "notes"}:
            width = 45
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    workbook.save(xlsx_path)
    return csv_path, xlsx_path, html_path
